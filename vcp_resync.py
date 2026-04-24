"""
VCP Data Sync: XLSX → SQL (Upsert Mode)
Reads the VCP Operations spreadsheet and generates SQL to upsert into Supabase.
Preserves existing data — adds new records and updates changed ones (status,
contact info, categories, match status, etc.). Does NOT wipe the database.

Usage:
  python vcp_resync.py "VCP Operations (DATA) (12).xlsx" > resync.sql

Split the output into 4 parts for Supabase SQL Editor (~100KB limit per paste):
  Part 1: Helper table + People upserts
  Part 2: Client application upserts
  Part 3: Coach application upserts
  Part 4: Match upserts + priority recalc + verification
"""

import openpyxl
import re
import sys


def esc(val):
    if val is None: return 'NULL'
    s = str(val).strip()
    if s == '' or s.lower() == 'none' or s == '#NAME?' or s == '#REF!': return 'NULL'
    s = s.replace("'", "''")
    return f"'{s}'"


def digits10(val):
    if val is None: return None
    s = re.sub(r'[^0-9]', '', str(val))[:10]
    return s if len(s) >= 7 else None


def norm_cohort(name):
    if not name: return None
    parts = str(name).strip().split()
    if len(parts) == 2: return f"{parts[0].capitalize()} {parts[1]}"
    return str(name).strip()


def map_client_status(desc):
    if not desc: return 'Waitlisted'
    s = str(desc).strip().lower()
    if s in ('active', 'a'): return 'Active'
    if s in ('completed', 'c'): return 'Completed'
    if s in ('dropped', 'd', 'client dropped'): return 'Dropped'
    if s in ('coach dropped', 'cd'): return 'Coach Dropped'
    if s in ('extended', 'e'): return 'Extended'
    if s in ('deferred',): return 'Deferred'
    if s in ('waitlisted', 'w', 'waitlist'): return 'Waitlisted'
    return 'Waitlisted'


def map_coach_status(desc):
    if not desc: return 'Waitlisted'
    s = str(desc).strip().lower()
    if s in ('active', 'a'): return 'Active'
    if s in ('completed', 'c'): return 'Completed'
    if 'dropped' in s: return 'Dropped'
    if s in ('waitlisted', 'w', 'waitlist'): return 'Waitlisted'
    if s in ('inactive',): return 'Inactive'
    if s in ('extended', 'e'): return 'Extended'
    return 'Waitlisted'


def map_match_status(desc):
    s = str(desc).strip().lower() if desc else ''
    if 'completed' in s: return 'Completed'
    if 'client dropped' in s or 'client drop' in s: return 'Dropped - Client'
    if 'coach dropped' in s or 'coach drop' in s: return 'Dropped - Coach'
    if 'dropped' in s: return 'Dropped - Mutual'
    if 'active' in s: return 'Active'
    return 'Active'


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else 'VCP Operations (DATA) (12).xlsx'
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    def safe(vals, idx):
        if idx < len(vals) and vals[idx] and str(vals[idx]).strip().lower() not in ('none', '', 'n/a'):
            return str(vals[idx]).strip()
        return None

    # ── PARSE CLIENTS (Masterlist Clients tab) ──
    ws_cl = wb['Masterlist Clients']
    clients = []
    for row in ws_cl.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals[0] or not str(vals[0]).strip(): continue
        client_id = digits10(vals[1])
        if not client_id: continue
        email = str(vals[5]).strip().rstrip(';').lower() if vals[5] and str(vals[5]).strip().lower() != 'none' else None
        clients.append({
            'cohort':       norm_cohort(vals[0]),
            'id':           client_id,
            'first':        str(vals[3]).strip().title() if vals[3] else 'Unknown',
            'last':         str(vals[4]).strip().title() if vals[4] else 'Unknown',
            'email':        email,
            'phone':        digits10(vals[7]),
            'branch':       str(vals[8]).strip() if vals[8] and str(vals[8]).strip().lower() != 'none' else None,
            'cat':          str(vals[2]).strip() if vals[2] and str(vals[2]).strip().lower() != 'none' else None,
            'waiver':       str(vals[9]).strip().lower() == 'true' if vals[9] else False,
            'coach_phone':  digits10(vals[13]),
            'status_desc':  str(vals[17]).strip() if vals[17] and str(vals[17]).strip().lower() != 'none' else None,
            'notes':        str(vals[15]).strip() if vals[15] and str(vals[15]).strip().lower() != 'none' else None,
            'prev':         str(vals[20]).strip().lower() == 'yes' if len(vals) > 20 and vals[20] else False,
        })

    # ── PARSE COACHES (Masterlist Coaches tab) ──
    ws_co = wb['Masterlist Coaches']
    coaches = []
    for row in ws_co.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals[0] or not str(vals[0]).strip(): continue
        coach_id = digits10(vals[4])
        if not coach_id: continue
        email = safe(vals, 7) if safe(vals, 7) else safe(vals, 10)
        if email: email = email.lower().rstrip(';')
        coaches.append({
            'cohort':      norm_cohort(vals[0]),
            'id':          coach_id,
            'first':       str(vals[7]).strip().title() if vals[7] else (str(vals[2]).strip().title() if vals[2] else 'Unknown'),
            'last':        str(vals[8]).strip().title() if vals[8] else (str(vals[3]).strip().title() if vals[3] else 'Unknown'),
            'email':       email,
            'phone':       digits10(vals[9]),
            'cat':         safe(vals, 5),
            'spec':        safe(vals, 6),
            'affiliation': safe(vals, 11),
            'profile':     safe(vals, 12),
            'status_desc': safe(vals, 19),
            'notes':       safe(vals, 22) if safe(vals, 22) and safe(vals, 22) != '.' else None,
            'prev':        str(vals[2]).strip().lower() == 'yes' if vals[2] and str(vals[2]).strip().lower() not in ('none', '#ref!') else False,
        })

    # ── DEDUP PEOPLE ──
    people = {}
    for c in clients:
        if c['id'] not in people:
            people[c['id']] = {'id': c['id'], 'type': 'client', 'first': c['first'], 'last': c['last'],
                               'email': c['email'], 'phone': c['phone'], 'branch': c['branch'], 'vet': True}
        elif c['email'] and not people[c['id']].get('email'):
            people[c['id']]['email'] = c['email']

    for co in coaches:
        if co['id'] not in people:
            people[co['id']] = {'id': co['id'], 'type': 'coach', 'first': co['first'], 'last': co['last'],
                                'email': co['email'], 'phone': co['phone'], 'branch': None, 'vet': False}
        elif co['email'] and not people[co['id']].get('email'):
            people[co['id']]['email'] = co['email']

    # ── BUILD MATCHES (coach_phone links client row to coach) ──
    coach_cohorts = set()
    for co in coaches:
        coach_cohorts.add((co['id'], co['cohort'].upper() if co['cohort'] else ''))

    match_set = {}
    for c in clients:
        if c['coach_phone'] and c['coach_phone'] in people:
            key = (c['cohort'].upper() if c['cohort'] else '', c['coach_phone'], c['id'])
            if key not in match_set:
                match_set[key] = {'cohort': c['cohort'], 'status': c['status_desc']}

    print(f'-- Parsed {len(clients)} client rows, {len(coaches)} coach rows')
    print(f'-- {len(people)} unique people, {len(match_set)} matches')
    print()
    print('-- ============================================================')
    print('-- VCP INCREMENTAL UPSERT (preserves existing data)')
    print('-- Run in Supabase SQL Editor, split into 4 parts if needed')
    print('-- ============================================================')
    print()

    # Helper table — recreated each run, does NOT affect real data
    print('-- Helper mapping table (dropped and recreated each run)')
    print('DROP TABLE IF EXISTS _pm;')
    print('CREATE TABLE _pm (lid TEXT PRIMARY KEY, pid UUID);')
    print()

    # ── STEP 1: UPSERT PEOPLE ──
    print('-- ============================================================')
    print('-- PART 1: Upsert people')
    print('-- ============================================================')
    print()
    for lid, p in people.items():
        print(f"DO $$ DECLARE v UUID; BEGIN")
        print(f"  INSERT INTO people (legacy_id,id_type,first_name,last_name,email,phone,is_veteran,branch_of_service)")
        print(f"  VALUES ({esc(p['id'])},{esc(p['type'])},{esc(p['first'])},{esc(p['last'])},{esc(p['email'])},{esc(p['phone'])},{str(p['vet']).lower()},{esc(p['branch'])})")
        print(f"  ON CONFLICT (legacy_id) DO UPDATE SET")
        print(f"    first_name        = COALESCE(EXCLUDED.first_name, people.first_name),")
        print(f"    last_name         = COALESCE(EXCLUDED.last_name, people.last_name),")
        print(f"    email             = COALESCE(EXCLUDED.email, people.email),")
        print(f"    phone             = COALESCE(EXCLUDED.phone, people.phone),")
        print(f"    branch_of_service = COALESCE(EXCLUDED.branch_of_service, people.branch_of_service),")
        print(f"    updated_at        = now()")
        print(f"  RETURNING id INTO v;")
        print(f"  IF v IS NULL THEN SELECT id INTO v FROM people WHERE legacy_id={esc(p['id'])}; END IF;")
        print(f"  INSERT INTO _pm VALUES ({esc(p['id'])},v) ON CONFLICT DO NOTHING;")
        print(f"END $$;")
        print()

    # ── STEP 2: UPSERT CLIENT APPLICATIONS ──
    print('-- ============================================================')
    print('-- PART 2: Upsert client applications')
    print('-- ============================================================')
    print()
    for c in clients:
        status = map_client_status(c['status_desc'])
        print(f"INSERT INTO client_applications")
        print(f"  (person_id,cohort_id,assigned_category,waiver_signed,status,status_notes,is_returning,participated_before)")
        print(f"VALUES (")
        print(f"  (SELECT pid FROM _pm WHERE lid={esc(c['id'])}),")
        print(f"  (SELECT id FROM cohorts WHERE cohort_name={esc(c['cohort'])}),")
        print(f"  {esc(c['cat'])},{str(c['waiver']).lower()},{esc(status)},{esc(c['notes'])},")
        print(f"  {str(c['prev']).lower()},{str(c['prev']).lower()}")
        print(f")")
        print(f"ON CONFLICT (person_id,cohort_id) DO UPDATE SET")
        print(f"  status             = EXCLUDED.status,")
        print(f"  status_notes       = EXCLUDED.status_notes,")
        print(f"  assigned_category  = COALESCE(EXCLUDED.assigned_category, client_applications.assigned_category),")
        print(f"  waiver_signed      = EXCLUDED.waiver_signed,")
        print(f"  is_returning       = EXCLUDED.is_returning,")
        print(f"  participated_before = EXCLUDED.participated_before,")
        print(f"  updated_at         = now();")
        print()

    # ── STEP 3: UPSERT COACH APPLICATIONS ──
    print('-- ============================================================')
    print('-- PART 3: Upsert coach applications')
    print('-- ============================================================')
    print()
    for co in coaches:
        status = map_coach_status(co['status_desc'])
        specs = co['spec']
        if specs:
            items = [s.strip() for s in specs.split(',') if s.strip()]
            arr = "ARRAY[" + ",".join(esc(s) for s in items) + "]::TEXT[]"
        else:
            arr = "ARRAY[]::TEXT[]"
        print(f"INSERT INTO coach_applications")
        print(f"  (person_id,cohort_id,coaching_specializations,assigned_category,affiliation,coach_profile_url,status,status_notes,is_returning)")
        print(f"VALUES (")
        print(f"  (SELECT pid FROM _pm WHERE lid={esc(co['id'])}),")
        print(f"  (SELECT id FROM cohorts WHERE cohort_name={esc(co['cohort'])}),")
        print(f"  {arr},{esc(co['cat'])},{esc(co['affiliation'])},{esc(co['profile'])},")
        print(f"  {esc(status)},{esc(co['notes'])},{str(co['prev']).lower()}")
        print(f")")
        print(f"ON CONFLICT (person_id,cohort_id) DO UPDATE SET")
        print(f"  status                  = EXCLUDED.status,")
        print(f"  status_notes            = EXCLUDED.status_notes,")
        print(f"  coaching_specializations = EXCLUDED.coaching_specializations,")
        print(f"  assigned_category       = COALESCE(EXCLUDED.assigned_category, coach_applications.assigned_category),")
        print(f"  affiliation             = COALESCE(EXCLUDED.affiliation, coach_applications.affiliation),")
        print(f"  coach_profile_url       = COALESCE(EXCLUDED.coach_profile_url, coach_applications.coach_profile_url),")
        print(f"  is_returning            = EXCLUDED.is_returning,")
        print(f"  updated_at              = now();")
        print()

    # Step 3b: auto-create missing coach apps for matched coaches not in coaches tab
    print('-- Step 3b: Auto-create missing coach applications for matched coaches')
    for c in clients:
        if not c['coach_phone'] or c['coach_phone'] not in people: continue
        cohort_key = (c['coach_phone'], c['cohort'].upper() if c['cohort'] else '')
        if cohort_key not in coach_cohorts:
            print(f"INSERT INTO coach_applications (person_id,cohort_id,status,is_returning)")
            print(f"SELECT p.id, c.id, 'Active', TRUE FROM people p, cohorts c")
            print(f"WHERE p.legacy_id={esc(c['coach_phone'])} AND c.cohort_name={esc(c['cohort'])}")
            print(f"ON CONFLICT (person_id,cohort_id) DO NOTHING;")
            print()

    # ── STEP 4: UPSERT MATCHES ──
    print('-- ============================================================')
    print('-- PART 4: Upsert matches + recalc priorities + verify')
    print('-- ============================================================')
    print()
    ok = 0
    for (coh_upper, coach_lid, client_lid), detail in match_set.items():
        if coach_lid not in people or client_lid not in people: continue
        mst = map_match_status(detail['status'])
        cn = detail['cohort']
        print(f"INSERT INTO new_matches (cohort_id,coach_application_id,client_application_id,status)")
        print(f"SELECT c.id, ka.id, ca.id, {esc(mst)}")
        print(f"FROM cohorts c")
        print(f"JOIN coach_applications ka ON ka.cohort_id=c.id")
        print(f"  AND ka.person_id=(SELECT pid FROM _pm WHERE lid={esc(coach_lid)})")
        print(f"JOIN client_applications ca ON ca.cohort_id=c.id")
        print(f"  AND ca.person_id=(SELECT pid FROM _pm WHERE lid={esc(client_lid)})")
        print(f"WHERE c.cohort_name={esc(cn)}")
        print(f"ON CONFLICT (client_application_id) DO UPDATE SET")
        print(f"  status     = EXCLUDED.status,")
        print(f"  updated_at = now();")
        print()
        ok += 1
    print(f'-- {ok} match upserts')
    print()

    # Recalc priorities
    print('-- Recalculate all priorities')
    print("SELECT recalc_all_priorities();")
    print()

    # Cleanup
    print('-- Cleanup helper table')
    print('DROP TABLE IF EXISTS _pm;')
    print()

    # Verification
    print("""-- Verification
SELECT 'People' AS tbl, COUNT(*) FROM people
UNION ALL SELECT 'Client Apps', COUNT(*) FROM client_applications
UNION ALL SELECT 'Coach Apps', COUNT(*) FROM coach_applications
UNION ALL SELECT 'Matches', COUNT(*) FROM new_matches;

SELECT c.cohort_name,
  (SELECT COUNT(*) FROM client_applications x WHERE x.cohort_id=c.id) AS clients,
  (SELECT COUNT(*) FROM coach_applications x WHERE x.cohort_id=c.id) AS coaches,
  (SELECT COUNT(*) FROM new_matches x WHERE x.cohort_id=c.id) AS matches
FROM cohorts c ORDER BY c.year, c.season;
""")

    wb.close()


if __name__ == '__main__':
    main()
